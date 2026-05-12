"""
Rebuild all three social calendars in OrangeBook-inspired voice.

Voice principles applied across all posts:
  - One central observation per post (not three or five)
  - Plain everyday language, sentences allowed to run long
  - No parallel triplets used for rhetorical impact
  - No "X is not Y, it is Z" reveal patterns
  - No punchy snap closes; trail off into the thought instead
  - No em dashes (use commas, periods, or "that is" instead)
  - Specific facts allowed to carry their own weight
  - Conversational register; sounds like someone thinking, not pronouncing

Three outputs:
  social/content-calendar-x.xlsx           - rewritten historical (X)
  social/content-calendar-threads.xlsx     - rewritten historical (Threads)
  social/content-calendar-metaphysics.xlsx - metaphysics, Threads-only, ~20 posts
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Shared styling
header_fill = PatternFill("solid", fgColor="1A5C3A")
header_font = Font(bold=True, color="FFFFFF")
alt_fill = PatternFill("solid", fgColor="FAF8F4")
thin = Side(style="thin", color="BBBBBB")
borders = Border(top=thin, bottom=thin, left=thin, right=thin)
center = Alignment(horizontal="center", vertical="top", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ===========================================================================
# HISTORICAL X (short form, ~200-275 chars)
# Discipline: pick one observation, state it plainly, stop.
# ===========================================================================

historical_x = [
    {
        "source": "Coins of the Caliphate",
        "post": ("The dinar has been in continuous use for over thirteen "
                 "hundred years. Caliph Abd al-Malik standardized it in "
                 "696 CE and eight national currencies still carry the "
                 "name today."),
    },
    {
        "source": "Risk and Rizq",
        "post": ("The English word hazard comes from al-zahr, the Arabic "
                 "word for dice. The Prophet prohibited transactions "
                 "where the buyer does not know what is being purchased. "
                 "Wall Street rediscovered the same problem in the 2000s."),
    },
    {
        "source": "The Word Tariff",
        "post": ("The word tariff comes from the Arabic ta'rif, meaning "
                 "to make known. It referred to customs schedules "
                 "posted in plain view so no merchant could be charged "
                 "different rates depending on who they were."),
    },
    {
        "source": "The Ottoman Experiment",
        "post": ("The Ottomans ran the largest economy of their time "
                 "for centuries without sovereign debt. In 1854 they "
                 "borrowed from European banks. Twenty seven years "
                 "later European officials were running their fiscal "
                 "administration."),
    },
    {
        "source": "The Sakk",
        "post": ("A merchant in 9th century Baghdad could write a sakk "
                 "and another merchant could cash it in Canton, four "
                 "thousand miles away. There were no banks. The "
                 "system ran on personal trust between agents who "
                 "knew each other by reputation."),
    },
    {
        "source": "The Waqf",
        "post": ("During the caliphate of Umar ibn Abd al-Aziz, zakat "
                 "was reportedly collected so effectively that "
                 "officials could not find eligible recipients. The "
                 "records are sparse but consistent on this point."),
    },
    {
        "source": "Forgotten Economists",
        "post": ("Ibn Khaldun wrote about supply and demand, the "
                 "division of labor, and the labor theory of value in "
                 "1377. The Muqaddimah predates Adam Smith by four "
                 "centuries. Smith almost certainly never read it."),
    },
    {
        "source": "Sukuk for Sovereigns",
        "post": ("The UK issued sovereign sukuk in 2014. Instead of "
                 "borrowing at interest, the government sold a share "
                 "of three buildings to investors and leased them "
                 "back. Whether this was substantively different from "
                 "a bond is the question Islamic finance has been "
                 "asking for forty years."),
    },
    {
        "source": "Marketplace of Medina",
        "post": ("The Prophet established a marketplace in Medina "
                 "where no one paid rent for a stall, no one hoarded "
                 "goods to manipulate prices, and no one sold what "
                 "they did not possess. The third rule alone would "
                 "rule out most modern derivatives."),
    },
    {
        "source": "Pakistan Transition",
        "post": ("Pakistan's Supreme Court has mandated the "
                 "elimination of riba from banking by January 2028. "
                 "Islamic banking is at 21.6 percent of assets. The "
                 "State Bank has no agreed replacement for the policy "
                 "interest rate."),
    },
    {
        "source": "The Funduk",
        "post": ("The funduk was a medieval trading complex that "
                 "combined a warehouse, a hotel, a currency exchange, "
                 "and a court for resolving disputes. The word became "
                 "fondaco in Italian and eventually fund in English."),
    },
    {
        "source": "Interest in Disguise",
        "post": ("In organized tawarruq a bank buys metal on the "
                 "London exchange, sells it to a customer at a "
                 "markup, and the customer sells it back at a small "
                 "loss. The metal never moves. The cash flow is "
                 "identical to a loan."),
    },
    {
        "source": "Petrodollars",
        "post": ("Gulf states have funded over a hundred and fifty "
                 "billion dollars of development aid through Islamic "
                 "institutions while parking reserves in interest-"
                 "bearing US Treasury bonds. The contradiction has "
                 "been around for decades."),
    },
    {
        "source": "Amir al-Bahr",
        "post": ("The word admiral comes from amir al-bahr, the "
                 "Arabic title for a commander of the sea. Fatimid "
                 "Egypt built the first professional navy in the "
                 "Mediterranean. European navies adopted the title "
                 "along with the structure."),
    },
    {
        "source": "Caravans to Cathay",
        "post": ("Mansa Musa's 1324 pilgrimage to Mecca brought "
                 "enough gold into Cairo that Egyptian prices stayed "
                 "elevated for the next decade. The Mamluk treasury "
                 "kept careful records of the disruption."),
    },
    {
        "source": "The Longest Verse",
        "post": ("The longest verse in the Quran is 2:282. The "
                 "subject is debt. Written contracts, two witnesses, "
                 "collateral for every loan. The subprime crisis "
                 "happened in part because lenders abandoned each "
                 "of these practices."),
    },
    {
        "source": "Before the Revelation",
        "post": ("Before Islam, a debtor who could not pay became "
                 "the creditor's property. Children were sometimes "
                 "sold to cover their parents' debts. The Quran "
                 "reversed this and made freeing a slave an "
                 "expiation for sins."),
    },
    {
        "source": "Ghazali's Warning",
        "post": ("Al-Ghazali argued in the 11th century that money "
                 "has no intrinsic value. It is a mirror reflecting "
                 "the value of other things. Selling the mirror to "
                 "buy another mirror was not, in his view, commerce."),
    },
    {
        "source": "Comparative",
        "post": ("Iran converted its banking system to Islamic "
                 "finance in 1983. Banks now charge fixed rates "
                 "approved by the central bank. Depositors never "
                 "lose money from bank losses. Whether this counts "
                 "as Islamic banking depends on what work the word "
                 "Islamic is doing."),
    },
    {
        "source": "Pakistan Transition",
        "post": ("Malaysia's Islamic banking reached 45.6 percent of "
                 "financing in 2023. It took forty years of "
                 "consistent policy. Pakistan's constitutional "
                 "deadline gives it less than two."),
    },
    {
        "source": "The Waqf",
        "post": ("By the Ottoman peak, somewhere between a quarter "
                 "and a third of productive land in the empire was "
                 "waqf property. The income funded hospitals, "
                 "universities, and public kitchens for centuries "
                 "with no government budget line."),
    },
    {
        "source": "Comparative",
        "post": ("The heter iska is a Jewish mechanism that "
                 "restructures a loan as a profit-sharing "
                 "partnership. The lender becomes a silent partner "
                 "sharing in gains and losses. The resemblance to "
                 "the Islamic mudarabah is probably not a coincidence."),
    },
    {
        "source": "Selling the Wind",
        "post": ("In 1637 Dutch traders sold tulip bulbs they did "
                 "not own at prices nobody could justify. The market "
                 "collapsed in days. Six centuries earlier the "
                 "Prophet had prohibited selling what you do not "
                 "possess."),
    },
    {
        "source": "After the Default",
        "post": ("Argentina has defaulted on its sovereign debt "
                 "nine times. The Quran encourages creditors to "
                 "forgive debts that cannot be repaid. Modern "
                 "finance calls this moral hazard. Which side is "
                 "creating the hazard is rarely asked."),
    },
    {
        "source": "New Contenders",
        "post": ("De-dollarization is mostly noise so far. The "
                 "dollar has structural advantages BRICS has not "
                 "begun to seriously challenge. China will not open "
                 "its capital account. The euro has no fiscal "
                 "union behind it."),
    },
    {
        "source": "House of Wisdom",
        "post": ("Cheques, bills of exchange, venture capital "
                 "partnerships, and correspondent banking all "
                 "existed in 9th century Baghdad. Much of the "
                 "financial infrastructure Europe later adopted was "
                 "developed there first."),
    },
    {
        "source": "Comparative",
        "post": ("Thomas Aquinas defended the ban on interest using "
                 "Aristotle, not just scripture. Money is consumed "
                 "in its use, so charging for its use charges for "
                 "something that does not separately exist. Calvin "
                 "rejected the argument in 1545."),
    },
    {
        "source": "Central Bank Question",
        "post": ("Every central bank manages the economy through "
                 "interest rates. Pakistan's deadline of January "
                 "2028 requires removing this tool. No Muslim "
                 "majority country has solved this problem at scale."),
    },
    {
        "source": "Karimi Merchants",
        "post": ("The Karimi merchants of medieval Egypt built "
                 "trade networks across the Indian Ocean that "
                 "lasted for centuries. Standardized contracts, "
                 "internal dispute resolution, profit-sharing "
                 "partnerships. No central authority involved."),
    },
    {
        "source": "Pakistan Transition",
        "post": ("Meezan Bank posted 101.5 billion rupees in profit "
                 "in 2024. Non-performing financing was 1.8 percent "
                 "against an industry average of 6.3 percent. "
                 "Whether it is structurally different from a "
                 "conventional bank is the harder question."),
    },
    {
        "source": "Comparative",
        "post": ("Calvin argued in 1545 that the Old Testament ban "
                 "on interest was specific to ancient Israel's "
                 "agrarian economy. Moderate interest in commerce, "
                 "he held, was permissible. Most of European "
                 "Protestantism followed him within a generation."),
    },
    {
        "source": "Amir al-Bahr",
        "post": ("The lateen sail was an Islamic innovation that "
                 "let ships sail closer to the wind than ancient "
                 "square sails permitted. It gave Muslim merchants "
                 "a structural advantage in Mediterranean trade for "
                 "roughly three centuries."),
    },
    {
        "source": "Ships and Storms",
        "post": ("A dhow leaving Aden for Gujarat in the 12th "
                 "century used mudarabah. The investor put up the "
                 "capital. The captain ran the voyage. They split "
                 "the profit if the ship made it home. Modern "
                 "venture capital rediscovered the structure eight "
                 "centuries later."),
    },
    {
        "source": "When Currencies Break",
        "post": ("Currency crises reveal something about governance. "
                 "External dollar pressure is roughly equal across "
                 "countries. What varies is who can protect ordinary "
                 "citizens from the consequences. Lebanon's banks "
                 "were captured by its own elites."),
    },
    {
        "source": "Forgotten Economists",
        "post": ("Abu Yusuf wrote in the 8th century that a just "
                 "state should run surpluses in good years and "
                 "spend them in bad ones. He grounded the argument "
                 "in the ruler's obligation to the governed. Keynes "
                 "rediscovered the policy in the 1930s."),
    },
    {
        "source": "Comparative",
        "post": ("The Franciscan school of the 1300s developed the "
                 "doctrine of lucrum cessans, compensation for "
                 "profit a lender would have earned elsewhere. Not "
                 "called interest. Functionally indistinguishable "
                 "from interest."),
    },
    {
        "source": "Pakistan Transition",
        "post": ("Pakistan issued over two trillion rupees in sukuk "
                 "in 2025. The government targets twenty percent "
                 "Shariah-compliant debt by FY28. Whether sukuk are "
                 "structurally different from bonds remains the open "
                 "question."),
    },
    {
        "source": "Debt Without Chains",
        "post": ("Global debt now exceeds three hundred trillion "
                 "dollars. The Quranic alternative is qard hasan: "
                 "interest-free loans returned when possible, "
                 "forgiven when not. The 2008 crisis cost the world "
                 "twenty-two trillion in lost output."),
    },
    {
        "source": "The Funduk",
        "post": ("The Arabic word makhzan was a warehouse where "
                 "goods were stored and catalogued. Venetian "
                 "merchants borrowed it as magazzino. It entered "
                 "French as magasin and English as magazine."),
    },
    {
        "source": "Building on Solid Ground",
        "post": ("In a conventional mortgage the bank is guaranteed "
                 "its return regardless of what happens to the "
                 "house. In a diminishing musharakah the bank "
                 "co-owns the house and shares the downside. The "
                 "2008 crisis showed which model passes the loss "
                 "onto the household."),
    },
    {
        "source": "The Dollar's World",
        "post": ("Reserve currency status is empire by other means. "
                 "The world prices oil in dollars, borrows in "
                 "dollars, and watches the Federal Reserve set the "
                 "cost of capital for everyone. Most of those "
                 "affected do not vote in American elections."),
    },
    {
        "source": "Marketplace of Medina",
        "post": ("The Prophet appointed a market inspector in "
                 "Medina in 622 CE. The position was called the "
                 "muhtasib. Checking weights, banning price "
                 "manipulation, protecting buyers from "
                 "misrepresentation. Western markets reinvented "
                 "these functions in the 19th and 20th centuries."),
    },
]


# ===========================================================================
# HISTORICAL THREADS (longer form, ~350-490 chars)
# Allowed to develop one idea more fully but still pick one focus.
# ===========================================================================

historical_threads = [
    {
        "source": "Coins of the Caliphate",
        "post": ("In 696 CE Caliph Abd al-Malik replaced the Byzantine "
                 "and Persian coinages then circulating across his "
                 "empire with a new gold dinar of standardized weight "
                 "and purity. Within a generation it became the most "
                 "trusted currency from Spain to China. Eight national "
                 "currencies still carry the name today, more than "
                 "thirteen hundred years later. Standardization and "
                 "consistent governance is what makes a currency "
                 "trusted, and the trust tends to outlast the empire."),
    },
    {
        "source": "Risk and Rizq",
        "post": ("Several modern financial words come directly from "
                 "Arabic. Hazard from al-zahr, the word for dice. Risk "
                 "from rizq, divine provision. Tariff from ta'rif, to "
                 "make known. Magazine from makhzan, a warehouse where "
                 "goods were catalogued. The vocabulary is not an "
                 "accident. Muslim merchants worked out the practical "
                 "problems of long distance commerce centuries before "
                 "Europeans did, and the vocabulary travelled along "
                 "with the techniques."),
    },
    {
        "source": "Ottoman Experiment",
        "post": ("The Ottomans ran the largest economy of their time "
                 "for centuries without sovereign debt. Cash waqfs "
                 "handled regulated lending. Wars were funded through "
                 "taxation. Public works ran on charitable endowments. "
                 "In 1854 they borrowed from European banks to fund "
                 "the Crimean War. By 1881 the Public Debt "
                 "Administration in Constantinople controlled a third "
                 "of state revenues with European officials in most "
                 "of the senior positions."),
    },
    {
        "source": "The Sakk",
        "post": ("A merchant in 9th century Baghdad could write a "
                 "sakk and another merchant could cash it in Canton, "
                 "four thousand miles away. No banks were involved in "
                 "the modern sense. The system ran on personal trust "
                 "between merchants who knew each other by reputation "
                 "and mutual obligation. The 2008 crisis was in part "
                 "a demonstration of what happens when the underlying "
                 "trust is replaced by complexity."),
    },
    {
        "source": "Forgotten Economists",
        "post": ("Ibn Khaldun wrote the Muqaddimah in 1377. The book "
                 "describes supply and demand, the division of labor, "
                 "the labor theory of value, and the rise and decline "
                 "of civilizations through productive labor and "
                 "consumption patterns. He wrote this four hundred "
                 "years before Adam Smith. Smith almost certainly "
                 "never read him, and the ideas had to be "
                 "rediscovered. The rediscovery tends to get the "
                 "credit."),
    },
    {
        "source": "Marketplace of Medina",
        "post": ("The Prophet established a marketplace in Medina in "
                 "622 CE with rules that look strikingly modern. No "
                 "rent for stalls, so any merchant could enter "
                 "regardless of wealth. No hoarding, enforced by a "
                 "market inspector called the muhtasib. No selling "
                 "what you do not possess, which would have ruled "
                 "out most modern derivatives. The Western world "
                 "rediscovered these ideas in the 19th and 20th "
                 "centuries and called them consumer protection and "
                 "securities regulation."),
    },
    {
        "source": "Pakistan Transition",
        "post": ("Pakistan's 26th Constitutional Amendment mandates "
                 "the elimination of riba by January 2028. Current "
                 "Islamic banking is at 21.6 percent of assets. Iran "
                 "tried full conversion in 1983 and ended up with a "
                 "system Islamic in legal form but conventional in "
                 "economic substance. Malaysia took forty years of "
                 "consistent policy to reach 45.6 percent. Whether "
                 "Pakistan's transition will be substantive or "
                 "nominal is the question the political conversation "
                 "tends to avoid."),
    },
    {
        "source": "Interest in Disguise",
        "post": ("In organized tawarruq a bank buys metal on the "
                 "London exchange, sells it to a customer at a "
                 "markup, and the customer immediately sells it back "
                 "at a small loss. The metal never moves out of the "
                 "warehouse. The cash flow is identical to an "
                 "interest-bearing loan. The Fiqh Academy has ruled "
                 "the practice impermissible. Many Islamic banks "
                 "continue to offer it. The gap between Islamic "
                 "finance theory and practice is where the hardest "
                 "questions live."),
    },
    {
        "source": "Petrodollars",
        "post": ("Gulf states have funded over a hundred and fifty "
                 "billion dollars of development aid through Islamic "
                 "institutions while parking petrodollar reserves in "
                 "interest-bearing US Treasury bonds. Both things "
                 "have been true for decades. The Islamic Development "
                 "Bank lends on concessional terms. The sovereign "
                 "wealth funds invest in whatever returns the most. "
                 "Acknowledging this would force a choice between "
                 "ideological consistency and financial pragmatism, "
                 "which is why it is rarely acknowledged."),
    },
    {
        "source": "Caravans to Cathay",
        "post": ("Mansa Musa's 1324 pilgrimage to Mecca brought "
                 "enough gold into Cairo that Egyptian prices stayed "
                 "elevated for the next decade. The Mamluk treasury "
                 "kept careful records of the disruption. His caravan "
                 "moved along trade routes that worked because "
                 "Islamic commercial law made contracts enforceable "
                 "across the empire and dispute resolution possible "
                 "across linguistic and political boundaries. The "
                 "legal infrastructure mattered more than the "
                 "geography."),
    },
    {
        "source": "The Longest Verse",
        "post": ("The longest verse in the Quran is 2:282. The subject "
                 "is debt. The verse requires written contracts, two "
                 "witnesses present at the transaction, collateral "
                 "for every loan, and specific documentation of "
                 "terms. The subprime mortgage crisis of 2008 "
                 "happened in large part because lenders had "
                 "abandoned each of these practices in pursuit of "
                 "volume. No documentation. No verified income. No "
                 "collateral worth what the loans were based on."),
    },
    {
        "source": "Before the Revelation",
        "post": ("Before Islam, a debtor who could not pay became "
                 "the creditor's property across much of the Near "
                 "East. Children were sometimes sold to cover their "
                 "parents' debts. The Quran reversed this entirely. "
                 "Freeing a slave became expiation for sins. "
                 "Compound interest was prohibited. Creditors were "
                 "instructed to give debtors time when they could "
                 "not pay, and to forgive entirely as a form of "
                 "charity. The reform restructured the relationship "
                 "between money and human freedom."),
    },
    {
        "source": "Comparative",
        "post": ("Iran converted its banking system to Islamic "
                 "finance in 1983. Forty years later banks charge "
                 "fixed rates approved in advance by the central "
                 "bank. These function identically to interest "
                 "rates from any external perspective. Depositors "
                 "have never lost money from bank losses because "
                 "the structure does not allow them to. The system "
                 "is Islamic by legal definition. By economic "
                 "substance it is conventional banking conducted in "
                 "Arabic terminology."),
    },
    {
        "source": "The Waqf",
        "post": ("By the Ottoman peak somewhere between a quarter "
                 "and a third of productive land in the empire was "
                 "waqf property. These were irrevocable charitable "
                 "endowments established by individual donors, not "
                 "government programs. Independent trustees managed "
                 "them according to the donor's deed. The "
                 "Suleymaniye Complex in Istanbul, Al-Azhar in "
                 "Cairo, and thousands of smaller institutions all "
                 "ran on this system. The modern state dismantled "
                 "most of it."),
    },
    {
        "source": "Ghazali's Warning",
        "post": ("Al-Ghazali argued in the 11th century that money "
                 "has no intrinsic value of its own. It is a mirror "
                 "reflecting the value of other things, and its "
                 "purpose is to facilitate exchange of real goods. "
                 "Selling the mirror to buy another mirror was not, "
                 "in his view, commerce. Modern derivatives markets "
                 "generate something like six hundred trillion "
                 "dollars of notional value annually, most of which "
                 "never connects to any real transaction. Al-Ghazali "
                 "would have recognized the problem immediately."),
    },
    {
        "source": "Selling the Wind",
        "post": ("In 1637 Dutch traders sold tulip bulbs they did "
                 "not own at prices nobody could justify. The market "
                 "collapsed in days. Six centuries earlier the "
                 "Prophet had prohibited bay' al-ma'dum, the sale "
                 "of something the seller does not possess. The "
                 "jurists who developed the doctrine were not being "
                 "anti-commercial. They were engineering a market "
                 "that structurally could not collapse this way. "
                 "Whether the constraint is worth the cost is a "
                 "fair argument."),
    },
    {
        "source": "After the Default",
        "post": ("Argentina has defaulted on its sovereign debt "
                 "nine times. Lebanon's banking system has "
                 "collapsed. Greece needed three bailouts. Pakistan "
                 "has been to the IMF twenty-three times. The "
                 "Quran encourages creditors to forgive debts they "
                 "know cannot be repaid. Modern finance calls this "
                 "moral hazard. Islamic economics calls persistent "
                 "lending to insolvent borrowers the deeper moral "
                 "hazard, since the lender is creating the "
                 "situation in the first place."),
    },
    {
        "source": "House of Wisdom",
        "post": ("Cheques, bills of exchange, venture capital "
                 "partnerships, and international correspondent "
                 "banking all existed in 9th century Baghdad. The "
                 "bayt al-mal was a state treasury. The mudarabah "
                 "was a venture capital partnership. The suftaja "
                 "was a bill of exchange. The Bait al-Hikma is "
                 "remembered as a translation center, and that is "
                 "true. It was also the financial capital of the "
                 "medieval world, and most of what Europe later "
                 "adopted in finance was developed there first."),
    },
    {
        "source": "Central Bank Question",
        "post": ("Every central bank manages its economy by "
                 "adjusting interest rates. The policy rate cools "
                 "inflation when raised and stimulates growth when "
                 "lowered. Remove interest rates and you remove the "
                 "primary tool of monetary policy. Pakistan faces "
                 "this with a constitutional deadline of January "
                 "2028. The honest answer is that nobody has built "
                 "a proven alternative for a three-hundred-billion-"
                 "dollar economy. The theoretical proposals exist. "
                 "The engineering work is largely undone."),
    },
    {
        "source": "Comparative",
        "post": ("Thomas Aquinas defended the medieval ban on "
                 "interest using Aristotle, not just scripture. The "
                 "argument was philosophical. Money is consumed in "
                 "its use, so charging for its use charges for "
                 "something that does not exist separately from the "
                 "money itself. Calvin rejected this in 1545 on the "
                 "grounds that the Old Testament ban was specific "
                 "to ancient Israel's agrarian economy. Within two "
                 "centuries interest had stopped being a sin in "
                 "most of European Christianity and started being "
                 "an industry."),
    },
    {
        "source": "Pakistan Transition",
        "post": ("Meezan Bank posted 101.5 billion rupees in profit "
                 "in 2024 with non-performing financing at 1.8 "
                 "percent against an industry average of 6.3 "
                 "percent. Deposits grew twenty-eight percent in "
                 "2025. Islamic banking is demonstrably viable as a "
                 "business. The harder question, which the industry "
                 "tends to avoid in public, is whether murabaha "
                 "financing is structurally different from a "
                 "conventional loan or simply relabeled. The cash "
                 "flows are often identical."),
    },
    {
        "source": "Amir al-Bahr",
        "post": ("The lateen sail was an Islamic innovation that "
                 "allowed ships to sail closer to the wind than the "
                 "square sails of antiquity permitted. It "
                 "transformed Mediterranean trade and gave Muslim "
                 "merchants a structural advantage for roughly "
                 "three centuries. The Portuguese eventually "
                 "learned the technique and used accumulated "
                 "Islamic nautical knowledge to enter the Indian "
                 "Ocean. The word admiral itself comes from amir "
                 "al-bahr."),
    },
    {
        "source": "Ships and Storms",
        "post": ("A dhow leaving Aden for Gujarat in the 12th "
                 "century carried pepper and silk. Pirates and "
                 "storms threatened the cargo. Insurance was "
                 "forbidden because it commodified uncertainty. "
                 "Interest was forbidden because money does not "
                 "create value alone. The solution was mudarabah. "
                 "The investor provided capital, the captain "
                 "expertise. If the ship sank the investor lost "
                 "money and the captain lost effort. If it returned "
                 "they split the profit. Venture capital "
                 "rediscovered this eight centuries later."),
    },
    {
        "source": "Forgotten Economists",
        "post": ("Abu Yusuf wrote in the 8th century that a just "
                 "state should run surpluses in good years and "
                 "spend them in bad ones. He grounded the "
                 "argument in the ruler's moral obligation to the "
                 "governed. The book was Kitab al-Kharaj, "
                 "commissioned by Caliph Harun al-Rashid. We call "
                 "the same policy countercyclical fiscal policy "
                 "now and attribute it to Keynes, who wrote about "
                 "twelve hundred years later."),
    },
    {
        "source": "Pakistan Transition",
        "post": ("Pakistan issued over two trillion rupees in "
                 "sukuk in 2025. Islamic government securities "
                 "grew from 12.6 percent of the total to 14.5 "
                 "percent in a single year. Green sukuk were five "
                 "to ten times oversubscribed. The demand for "
                 "Shariah-compliant sovereign debt clearly exists "
                 "and the supply is catching up. Whether sukuk are "
                 "structurally different from conventional bonds "
                 "in their economic effects is the question the "
                 "government conversation does not seem in a "
                 "hurry to settle."),
    },
    {
        "source": "Debt Without Chains",
        "post": ("Global debt now exceeds three hundred trillion "
                 "dollars. The Quranic alternative is qard hasan, "
                 "an interest-free loan returned at principal when "
                 "the borrower is able and forgiven when they "
                 "cannot pay. It sounds naive in a world of "
                 "leveraged buyouts and structured products. The "
                 "2008 crisis cost the world economy roughly "
                 "twenty-two trillion dollars in lost output. The "
                 "question worth asking is whether the current "
                 "system is sustainable on its own terms."),
    },
    {
        "source": "The Funduk",
        "post": ("The Arabic word makhzan referred to a warehouse "
                 "where goods were stored and catalogued. Venetian "
                 "merchants borrowed it as magazzino. It entered "
                 "French as magasin and English as magazine. The "
                 "same thing happened with funduk (Italian fondaco), "
                 "sakk (cheque), suq (souk), and many others. "
                 "Medieval Islamic commerce gave the European "
                 "trading system most of its vocabulary and a fair "
                 "amount of its infrastructure."),
    },
    {
        "source": "Building on Solid Ground",
        "post": ("In a conventional mortgage the bank is "
                 "guaranteed its return regardless of what happens "
                 "to the property's value. The risk of decline "
                 "sits entirely with the borrower. In a diminishing "
                 "musharakah the bank co-owns the house and shares "
                 "the risk of its value falling. If prices "
                 "collapse, both parties absorb part of the loss. "
                 "The 2008 housing crisis demonstrated which model "
                 "passes the cost of failure onto the household "
                 "that could least afford it."),
    },
]


# ===========================================================================
# METAPHYSICS THREADS (~450-490 chars, ~20 posts)
# ===========================================================================

metaphysics_threads = [
    # Historical quotes
    {
        "category": "Historical Quote",
        "post": ("Ibn Arabi, who died in 1240, described reality this "
                 "way. The whole world is nothing but the self-"
                 "disclosure of the Real through the divine names. "
                 "This is what classical Islam meant by economic "
                 "provision. Not income. Not output. A tajalli, a "
                 "self-disclosure of divine action. The shift in "
                 "modern Islamic economics from this view to a "
                 "modeling view where rizq becomes an income variable "
                 "may be the central change in the field. It is "
                 "rarely discussed as such."),
        "notes": "Source: Futuhat al-Makkiyya, via Chittick (1989)",
    },
    {
        "category": "Historical Quote",
        "post": ("Quran 6:59 describes the structure of reality "
                 "directly. With Him are the keys of the Unseen. "
                 "None knows them but He. He knows what is in the "
                 "land and the sea. Not a leaf falls but that He "
                 "knows it. This is a two-layer ontology. A "
                 "Manifest layer that economics has traditionally "
                 "worked on, and an Unseen layer where rizq and "
                 "qadar are determined before they appear in time. "
                 "The relation between the two is what the "
                 "metaphysical concepts are about."),
        "notes": "Quran 6:59",
    },
    {
        "category": "Historical Quote",
        "post": ("Mulla Sadra wrote in the 17th century that "
                 "existence is the principle and quiddity is the "
                 "manifestation. In plain English, being itself "
                 "comes first. The kinds of things that exist are "
                 "how being shows up. Contemporary philosophy of "
                 "physics has been groping toward something similar "
                 "since at least the 1970s. Sadra got there by way "
                 "of the Akbarian tradition and a careful reading "
                 "of Ibn Sina. Worth reading him in Izutsu if Asfar "
                 "itself is out of reach."),
        "notes": "Source: Asfar al-Arba'a; cf Izutsu.",
    },
    {
        "category": "Historical Quote",
        "post": ("Shah Wali Allah of Delhi wrote in the 18th "
                 "century that the names of God are the principles "
                 "by which the manifest order is governed. Economic "
                 "regularities, on this view, are not laws of "
                 "matter operating independently. They are patterns "
                 "of how divine attributes self-disclose in "
                 "spacetime. The framework is wider than modern "
                 "materialism allows. Hermansen's translation of "
                 "Hujjat Allah al-Baligha is the place to read him "
                 "in English."),
        "notes": "Cite Hermansen translation.",
    },

    # Methodology
    {
        "category": "Methodology",
        "post": ("There are two kinds of question one can ask about "
                 "Islamic economics. One is how rizq distributes "
                 "across households or how Islamic markets behave "
                 "under certain constraints. Economics has methods "
                 "for these. The other question is what rizq is at "
                 "the level of being. Different methods, different "
                 "answer. Most contemporary work has answered the "
                 "first. The classical tradition was working on the "
                 "second. The gap between them is the project worth "
                 "doing."),
        "notes": "(B)-not-(A) distinction.",
    },
    {
        "category": "Methodology",
        "post": ("The argument that quantum mechanics proves "
                 "classical Islamic doctrines is not worth defending. "
                 "Physics does not prove ontologies. What modern "
                 "physics has done, through Bell's theorem, the "
                 "measurement problem, and the Penrose-Godel "
                 "argument, is to undermine the picture of reality "
                 "that 19th century materialism took for granted. "
                 "That picture had foreclosed certain metaphysical "
                 "questions. With it gone, the questions are "
                 "askable again. The doctrines then have to make "
                 "their own case."),
        "notes": "(B)-not-(A) discipline.",
    },
    {
        "category": "Methodology",
        "post": ("If Penrose's argument from Godel's incompleteness "
                 "theorem is right, and the mathematics is settled "
                 "even if the interpretation is debated, then "
                 "human mathematical understanding cannot be "
                 "captured by any algorithmic procedure. The "
                 "implication for Islamic economics is that niyyah, "
                 "the orientation of the soul, cannot be a "
                 "computational state. This becomes more relevant "
                 "as AI systems get better at mimicking human "
                 "output without doing anything resembling "
                 "intention."),
        "notes": "Penrose-Godel for niyyah.",
    },
    {
        "category": "Methodology",
        "post": ("There are productive lines of work in heterodox "
                 "Islamic economics. Ergodicity applied to riba "
                 "dynamics. Network theory applied to family "
                 "provision. Behavioral mediation applied to the "
                 "effects of gratitude and tawakkul. All of this "
                 "is valuable and none of it is metaphysics. The "
                 "applied work models dynamics in the manifest "
                 "layer. It does not address what the phenomena "
                 "are at the level of being. Separating the two "
                 "registers lets each be itself."),
        "notes": "Heterodox vs metaphysics.",
    },

    # Reading recommendations
    {
        "category": "Reading Rec",
        "post": ("If you want to read Akbarian metaphysics "
                 "seriously in English, the place to start is "
                 "William Chittick's Sufi Path of Knowledge. It is "
                 "the standard scholarly work on Ibn Arabi. Read "
                 "it alongside Chittick's Self-Disclosure of God, "
                 "the more systematic companion volume. Sachiko "
                 "Murata's Tao of Islam gives an alternative angle "
                 "through comparative metaphysics using Chinese "
                 "categories. These three together are the "
                 "indigenous-language core. There is no shortcut."),
        "notes": "SUNY Press; AbeBooks for used copies.",
    },
    {
        "category": "Reading Rec",
        "post": ("For the philosophical implications of quantum "
                 "mechanics, David Albert's Quantum Mechanics and "
                 "Experience is the right entry point. Albert is "
                 "both a philosopher and a trained physicist, "
                 "which is unusual and useful. The book is short, "
                 "around two hundred pages, and technically "
                 "careful. It treats the measurement problem "
                 "head-on rather than around it. Pair with Tim "
                 "Maudlin's Philosophy of Physics, Quantum Theory "
                 "for a more recent complement."),
        "notes": "Harvard University Press.",
    },
    {
        "category": "Reading Rec",
        "post": ("Peter Adamson's Philosophy in the Islamic World, "
                 "volume three of his History of Philosophy Without "
                 "Any Gaps, is the best modern overview of the "
                 "Islamic philosophical tradition in English. It "
                 "reads like a long accessible essay rather than a "
                 "textbook, which is unusual for the genre. "
                 "Coverage runs from al-Kindi through Ibn Arabi to "
                 "Mulla Sadra and the Indian school. The companion "
                 "podcast is also free and useful for orientation."),
        "notes": "Oxford University Press.",
    },

    # Open questions
    {
        "category": "Open Question",
        "post": ("What is rizq? Two readings are common. The "
                 "economic reading treats it as the income or "
                 "output a household receives in a given period. "
                 "Most modern Islamic economics has worked with "
                 "this. The metaphysical reading treats it as the "
                 "actualization into spacetime of provision "
                 "determined in the Unseen. The classical "
                 "tradition was working on this. The shift from "
                 "the second to the first happened quietly over "
                 "the second half of the 20th century, and the "
                 "field has not entirely recovered from it."),
        "notes": "Rizq as actualization.",
    },
    {
        "category": "Open Question",
        "post": ("What is barakah? One reading treats it as a "
                 "multiplicative coefficient on wealth dynamics, "
                 "where your money or time goes further than the "
                 "nominal magnitude would predict. Heterodox "
                 "economics can model this. The other reading "
                 "treats it as the thickness of manifestation. A "
                 "measure of how much of the underlying field of "
                 "divine names is disclosed in a particular event. "
                 "Both readings can be coherent. They answer "
                 "different questions about different objects."),
        "notes": "Barakah as coupling.",
    },
    {
        "category": "Open Question",
        "post": ("There are two readings of riba in the literature. "
                 "The narrower one treats it as a juridical "
                 "prohibition on certain contract forms. Most "
                 "Islamic finance practice is built on this. The "
                 "wider reading treats riba as ontological "
                 "inversion. The claim that capital can grow "
                 "autonomously without coupling to productive "
                 "engagement with reality, attributing creative "
                 "power to a created thing. The wider reading "
                 "applies to AI-managed trading as directly as to "
                 "medieval moneylending."),
        "notes": "Riba as ontological inversion.",
    },

    # Anti-mysticism
    {
        "category": "Anti-Mysticism",
        "post": ("The claim that quantum mechanics proves God is "
                 "bad on both sides. It is bad epistemology because "
                 "physics does not prove metaphysical claims. It "
                 "constrains the space of possible ontologies, and "
                 "constraints are not proofs. It is bad theology "
                 "because tying religious truth to a physics theory "
                 "makes religion hostage to that theory's fate. "
                 "What physics actually does for metaphysics is "
                 "narrower than the popular claim, and the narrower "
                 "thing is more interesting."),
        "notes": "Public statement of (B)-not-(A).",
    },
    {
        "category": "Anti-Mysticism",
        "post": ("There are essentially two literatures on quantum "
                 "mechanics and religion. The popular literature "
                 "includes Capra's Tao of Physics, Chopra's books, "
                 "and a steady supply of YouTube essays. This uses "
                 "quantum as a license for almost any spiritual "
                 "claim and overreaches consistently. The academic "
                 "literature includes Maudlin and Albert on "
                 "philosophy of physics, and journals like "
                 "Foundations of Physics. Technical, modest in what "
                 "it claims, and almost completely absent from the "
                 "public conversation."),
        "notes": "Pop-mysticism vs serious work.",
    },

    # Physics notes
    {
        "category": "Physics Note",
        "post": ("Leonard Susskind, one of the founders of string "
                 "theory and a central figure in the holographic "
                 "principle, said something striking in a recent "
                 "interview. We live in the wrong kind of world to "
                 "be described by string theory. I can tell you "
                 "with absolute certainty that it is not the real "
                 "world we live in. We need to start over. The "
                 "state of theoretical physics is more open than "
                 "the popular literature suggests, and this matters "
                 "for any project that wants to use holography for "
                 "metaphysical purposes."),
        "notes": "Jaimungal interview, 2026.",
    },
    {
        "category": "Physics Note",
        "post": ("Bell's theorem gets invoked vaguely in popular "
                 "writing, so it is worth being precise. Bell "
                 "proved in 1964 a mathematical inequality that "
                 "any theory must satisfy if physical systems "
                 "have definite properties before measurement and "
                 "interactions propagate locally through space. "
                 "The 2015 loophole-free experiments by Hensen "
                 "and colleagues showed the inequality is "
                 "violated. Physical reality cannot have both "
                 "properties. One has to go, and this is forced "
                 "by experiment."),
        "notes": "Hensen et al. 2015.",
    },

    # Public discourse
    {
        "category": "Public Discourse",
        "post": ("On classical Islamic miracles and physics, the "
                 "disciplined position is narrow. The event "
                 "happened per the textual tradition. Current "
                 "physics has no mechanism for it. Whether future "
                 "physics will, or whether the event operates at "
                 "a level physics describes only partially, is "
                 "genuinely open. The theological claim does not "
                 "depend on either resolution. Anything stronger "
                 "in either direction overreaches. The honest "
                 "posture is curiosity, not premature closure."),
        "notes": "Physics-open agnosticism.",
    },
    {
        "category": "Public Discourse",
        "post": ("If AI restructures the material economy as some "
                 "predict, the applied-economics frameworks that "
                 "depend on current market structure will need "
                 "substantial rebuilding. Income variance, "
                 "household provisioning, behavioral feedback in "
                 "conventional settings. The metaphysical "
                 "foundations of Islamic economic thought are "
                 "differently situated. What rizq and barakah and "
                 "qadar are at the level of being. These are "
                 "structural claims, not regularity claims, and "
                 "material restructuring does not touch them."),
        "notes": "Engages Falk-Tsoukalas paper.",
    },
]


# ===========================================================================
# Build helper
# ===========================================================================

header_fill = PatternFill("solid", fgColor="1A5C3A")
header_font = Font(bold=True, color="FFFFFF")
alt_fill = PatternFill("solid", fgColor="FAF8F4")
thin = Side(style="thin", color="BBBBBB")
borders = Border(top=thin, bottom=thin, left=thin, right=thin)
center = Alignment(horizontal="center", vertical="top", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)


def write_sheet(wb, name, rows, columns_spec):
    ws = wb.create_sheet(name)
    headers, widths = zip(*columns_spec)
    ws.append(list(headers))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = borders
    for idx, row_vals in enumerate(rows, start=2):
        ws.append(row_vals)
        row = ws[idx]
        if idx % 2 == 0:
            for c in row:
                c.fill = alt_fill
        for c in row:
            c.border = borders
            c.alignment = left_top
        ws.row_dimensions[idx].height = 90
    ws.freeze_panes = "A2"


def build_historical_x():
    wb = Workbook()
    wb.remove(wb.active)
    cols = [
        ("Date", 12), ("Time", 10), ("Post", 70), ("Type", 14),
        ("Thread_ID", 12), ("Chars", 8), ("Source Episode", 22), ("Status", 12),
    ]
    rows = [
        ["", "", p["post"], "standalone", "", len(p["post"]), p["source"], "draft"]
        for p in historical_x
    ]
    write_sheet(wb, "X Content Calendar", rows, cols)
    out = "/Users/hamzazahid/Library/CloudStorage/GoogleDrive-88hzahid@gmail.com/My Drive/IslamicEconomics/social/content-calendar-x.xlsx"
    wb.save(out)
    return out, len(rows)


def build_historical_threads():
    wb = Workbook()
    wb.remove(wb.active)
    cols = [
        ("Date", 12), ("Time", 10), ("Post", 70), ("Type", 14),
        ("Thread_ID", 12), ("Chars", 8), ("Source Episode", 22), ("Status", 12),
    ]
    rows = [
        ["", "", p["post"], "standalone", "", len(p["post"]), p["source"], "draft"]
        for p in historical_threads
    ]
    write_sheet(wb, "Threads Content Calendar", rows, cols)
    out = "/Users/hamzazahid/Library/CloudStorage/GoogleDrive-88hzahid@gmail.com/My Drive/IslamicEconomics/social/content-calendar-threads.xlsx"
    wb.save(out)
    return out, len(rows)


def build_metaphysics():
    wb = Workbook()
    wb.remove(wb.active)
    cols = [
        ("Date", 12), ("Time", 10), ("Post", 70), ("Category", 18),
        ("Type", 14), ("Chars", 8), ("Notes", 30), ("Status", 12),
    ]
    rows = [
        ["", "", p["post"], p["category"], "standalone", len(p["post"]),
         p.get("notes", ""), "draft"]
        for p in metaphysics_threads
    ]
    write_sheet(wb, "Metaphysics Threads", rows, cols)
    out = "/Users/hamzazahid/Library/CloudStorage/GoogleDrive-88hzahid@gmail.com/My Drive/IslamicEconomics/social/content-calendar-metaphysics.xlsx"
    wb.save(out)
    return out, len(rows)


if __name__ == "__main__":
    for fn in (build_historical_x, build_historical_threads, build_metaphysics):
        path, n = fn()
        print(f"Wrote {n} posts to {path}")

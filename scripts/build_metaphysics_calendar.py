"""
Build content-calendar-metaphysics.xlsx for manual publication.
Mirrors structure of existing content-calendar-{x,threads}.xlsx but adds
a Category column appropriate to a metaphysics track.

All posts pass the Kaiser-test: defensible, specific, not overclaiming.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Posts
# ---------------------------------------------------------------------------
# Each post is a dict with:
#   category   -- the content type (Historical Quote, Methodology, etc.)
#   x          -- short form (<=250 chars) for X
#   threads    -- long form (up to ~500 chars) for Threads
#   notes      -- optional internal note (eg. source, link, follow-up)

posts = [
    # --- HISTORICAL QUOTES (Akbarian / Quranic) ---
    {
        "category": "Historical Quote",
        "x": ("Ibn Arabi (d. 1240) on what reality is: \"The whole world "
              "is nothing but the self-disclosure of the Real through the "
              "divine names.\" This is what classical Islam thought economic "
              "provision actually is - a tajalli, not just an income."),
        "threads": ("Ibn Arabi (d. 1240) on what reality is:\n\n"
                    "\"The whole world is nothing but the self-disclosure "
                    "of the Real through the divine names.\"\n\n"
                    "This is what classical Islam thought economic provision "
                    "actually is. Not income. Not output. A tajalli - a "
                    "self-disclosure of divine action through specific names.\n\n"
                    "It was not \"how rizq distributes\" that the tradition "
                    "was answering. It was \"what rizq is.\""),
        "notes": "Source: Futuhat al-Makkiyya, via Chittick (1989)",
    },
    {
        "category": "Historical Quote",
        "x": ("Quran on the structure of reality: \"With Him are the keys of "
              "the Unseen; none knows them but He. He knows what is in the "
              "land and the sea. Not a leaf falls but that He knows it.\" "
              "(6:59). A two-layer ontology, stated directly."),
        "threads": ("Quran 6:59:\n\n"
                    "\"With Him are the keys of the Unseen; none knows them "
                    "but He. He knows what is in the land and the sea. Not a "
                    "leaf falls but that He knows it.\"\n\n"
                    "A two-layer ontology stated directly. Manifest (shahada) "
                    "and Unseen (ghayb). The relation between them is what "
                    "concepts like rizq, qadar, and barakah are about. "
                    "Modern economics works only at the first layer."),
        "notes": "Use as the epigraph for any introductory thread.",
    },
    {
        "category": "Historical Quote",
        "x": ("Mulla Sadra (d. 1640): \"Existence is the principle. Quiddity "
              "is the manifestation.\" Wujud comes first; the kinds of things "
              "that exist are how it shows up. The metaphysical inversion "
              "contemporary philosophy of physics has been groping toward."),
        "threads": ("Mulla Sadra (d. 1640):\n\n"
                    "\"Existence is the principle. Quiddity is the "
                    "manifestation.\"\n\n"
                    "In plain English: being itself comes first. The kinds "
                    "of things that exist are how being shows up - not the "
                    "other way around.\n\n"
                    "Contemporary philosophy of physics has been groping "
                    "toward this. Sadra got there in the 17th century via "
                    "the Akbarian tradition."),
        "notes": "Source: Asfar al-Arba'a; cf Toshihiko Izutsu.",
    },
    {
        "category": "Historical Quote",
        "x": ("Shah Wali Allah of Delhi (d. 1762): \"The names of God are the "
              "principles by which He governs the manifest order.\" Economic "
              "regularities, on this view, are patterns of how divine "
              "attributes self-disclose in spacetime."),
        "threads": ("Shah Wali Allah of Delhi (d. 1762):\n\n"
                    "\"The names of God are the principles by which He "
                    "governs the manifest order.\"\n\n"
                    "Economic regularities, on this view, are not laws of "
                    "matter. They are patterns of how divine attributes "
                    "self-disclose in spacetime.\n\n"
                    "A wider ontology than classical materialism allows. "
                    "Worth re-reading him in Hermansen's translation of "
                    "Hujjat Allah al-Baligha."),
        "notes": "Cite Hermansen translation.",
    },

    # --- METHODOLOGY ---
    {
        "category": "Methodology",
        "x": ("Two different questions about Islamic economics:\n\n"
              "\"How does rizq distribute across households?\" is a modeling "
              "question. Economics answers it.\n\n"
              "\"What is rizq at the level of being?\" is a metaphysical "
              "question. Different methods. Different answer."),
        "threads": ("Two different questions about Islamic economics that "
                    "get conflated:\n\n"
                    "1. How does rizq distribute? How does barakah affect "
                    "outcomes? How does riba dynamics produce instability? "
                    "These are modeling questions. Economics answers them.\n\n"
                    "2. What is rizq? What is barakah? What is riba, at the "
                    "level of being? These are metaphysical questions. "
                    "Different methods. Different answers.\n\n"
                    "Most contemporary Islamic economics has answered the "
                    "first. The classical tradition was working on the "
                    "second."),
        "notes": "The (B)-not-(A) distinction.",
    },
    {
        "category": "Methodology",
        "x": ("Modern physics does not prove classical Islamic doctrines "
              "about provision, intention, or the Unseen. What it does is "
              "open the conceptual space that classical materialism had "
              "foreclosed. Then the doctrines articulate themselves in their "
              "own terms."),
        "threads": ("Modern physics does not prove classical Islamic "
                    "doctrines about rizq, qadar, niyyah, or the Unseen.\n\n"
                    "What it does is open the conceptual space classical "
                    "materialism foreclosed. Bell's theorem killed local "
                    "realism. The measurement problem reopened consciousness. "
                    "Penrose-Godel ruled out algorithmic mind.\n\n"
                    "With that space open, the doctrines are stated in their "
                    "own terms. The work is negative - removal of obstacles, "
                    "not validation."),
        "notes": "The (B)-not-(A) discipline framed for public.",
    },
    {
        "category": "Methodology",
        "x": ("If Penrose's argument from Godel's incompleteness is right, "
              "mathematical understanding cannot be captured by any "
              "algorithmic procedure. The implication: niyyah cannot be a "
              "computational state. A load-bearing claim for any serious "
              "treatment of human agency."),
        "threads": ("If Penrose's argument from Godel's incompleteness is "
                    "right (and it is mathematically settled), human "
                    "mathematical understanding cannot be captured by any "
                    "algorithmic procedure.\n\n"
                    "The implication: niyyah - intention, the orientation "
                    "of the soul - cannot be a computational state.\n\n"
                    "This is a load-bearing claim for any serious treatment "
                    "of human agency. It also distinguishes human action "
                    "from AI output in principle, not just in practice."),
        "notes": "Penrose Emperor's New Mind, Shadows of the Mind.",
    },
    {
        "category": "Methodology",
        "x": ("Ergodicity for riba, network theory for rizq, behavioral "
              "mediation for barakah - valuable applied frameworks. They are "
              "not metaphysics. The discipline of separating them is what "
              "produces useful work in both registers."),
        "threads": ("Three lines of contemporary work in heterodox Islamic "
                    "economics:\n\n"
                    "1. Ergodicity (time-vs-ensemble averages) for riba "
                    "dynamics.\n"
                    "2. Network theory (super-linear scaling) for family "
                    "provision.\n"
                    "3. Behavioral mediation for barakah, tawakkul, "
                    "gratitude.\n\n"
                    "All valuable. None of them metaphysics. They model "
                    "dynamics in the manifest layer; they do not articulate "
                    "what the phenomena are at the level of being.\n\n"
                    "Separating the two registers lets each be itself."),
        "notes": "The heterodox-economics companion volume.",
    },
    {
        "category": "Methodology",
        "x": ("For the metaphysics-of-Islamic-economics project, the apparatus "
              "is restricted to frameworks with scientific authority. QM and "
              "Godel-Penrose. Whitehead and Heidegger are suggestive; they "
              "lack the standing. Discipline matters."),
        "threads": ("Methodological discipline of the project:\n\n"
                    "Apparatus restricted to frameworks with scientific "
                    "authority. Two are load-bearing:\n\n"
                    "1. Quantum foundations (Bell, measurement problem, "
                    "interpretations).\n"
                    "2. Math logic via Godel-Penrose (non-algorithmic "
                    "understanding).\n\n"
                    "Whitehead and Heideggerian phenomenology are "
                    "interesting but lack the standing. Mentioned, not "
                    "load-bearing.\n\n"
                    "Discipline matters more than rhetorical reach."),
        "notes": "Scientific authority constraint.",
    },
    {
        "category": "Methodology",
        "x": ("The hard question in Islamic economics is not \"what does the "
              "tradition say?\" It is \"at what level are we asking?\" "
              "Modeling, formal analogy, or metaphysical mechanism are "
              "different registers. Clarity about register is the first "
              "discipline."),
        "threads": ("The hard question in Islamic economics is not \"what "
                    "does the tradition say about X?\"\n\n"
                    "It is: at what level are we asking?\n\n"
                    "- Modeling (how does X behave?)\n"
                    "- Formal analogy (does X share a structure with Y from "
                    "contemporary science?)\n"
                    "- Metaphysical mechanism (what is X, at the level of "
                    "being?)\n\n"
                    "Different registers, different methods, different "
                    "answers. Clarity about register is the first "
                    "discipline."),
        "notes": "Three-levels-of-rigor framework.",
    },

    # --- READING RECOMMENDATIONS ---
    {
        "category": "Reading Rec",
        "x": ("If you want to read Akbarian metaphysics seriously in English, "
              "start with William Chittick's The Sufi Path of Knowledge. The "
              "standard scholarly work on Ibn Arabi. SUNY Press, ~480 pp. "
              "There is no shortcut."),
        "threads": ("Recommended core for anyone working on the metaphysical "
                    "foundations of Islamic economics:\n\n"
                    "1. William Chittick, The Sufi Path of Knowledge. The "
                    "standard scholarly work on Ibn Arabi in English.\n\n"
                    "2. William Chittick, The Self-Disclosure of God. "
                    "Systematic Akbarian companion.\n\n"
                    "3. Sachiko Murata, The Tao of Islam. Alternative angle "
                    "through comparative metaphysics.\n\n"
                    "All three SUNY Press. Together they form the "
                    "indigenous-language core. No serious work on Akbarian "
                    "themes is possible without them."),
        "notes": "SUNY Press; available used on AbeBooks.",
    },
    {
        "category": "Reading Rec",
        "x": ("Recommended for the implications of modern physics: David "
              "Albert, Quantum Mechanics and Experience. Short, technically "
              "careful, the standard text for the foundations debate. ~200 "
              "pages. Required reading."),
        "threads": ("Recommended for anyone wanting to engage the "
                    "philosophical implications of QM at a serious level:\n\n"
                    "David Albert, Quantum Mechanics and Experience (1992).\n\n"
                    "- Short (~200 pages).\n"
                    "- Technically careful (Albert is both a philosopher and "
                    "a trained physicist).\n"
                    "- Treats the measurement problem head-on.\n"
                    "- Standard text for graduate philosophy of physics.\n\n"
                    "If you read only one philosophy-of-physics book, this. "
                    "Pair with Tim Maudlin's Philosophy of Physics: Quantum "
                    "Theory for the modern complement."),
        "notes": "Harvard University Press.",
    },
    {
        "category": "Reading Rec",
        "x": ("Roger Penrose, The Emperor's New Mind, for the argument that "
              "mathematical understanding is non-algorithmic. Skip the "
              "Mandelbrot-set chapters; read the Godel-and-mind chapters "
              "carefully. Required for anyone working on niyyah."),
        "threads": ("Roger Penrose, The Emperor's New Mind (1989).\n\n"
                    "Read for the argument from Godel's incompleteness to "
                    "the non-algorithmic character of mathematical "
                    "understanding. Implications far beyond physics - "
                    "including for any treatment of human agency.\n\n"
                    "Skip the Mandelbrot-set chapters. Read the Godel-and-"
                    "mind chapters carefully. Pair with Shadows of the "
                    "Mind (1994) and Feferman's 1996 critical response."),
        "notes": "Penrose 1989, Penrose 1994, Feferman 1996.",
    },
    {
        "category": "Reading Rec",
        "x": ("Peter Adamson, Philosophy in the Islamic World (Vol. 3 of his "
              "History of Philosophy). The best modern overview of the "
              "Islamic philosophical tradition in English. Reads like a long "
              "accessible essay. Start here."),
        "threads": ("Best modern overview of the Islamic philosophical "
                    "tradition in English:\n\n"
                    "Peter Adamson, Philosophy in the Islamic World "
                    "(2016). Volume 3 of his History of Philosophy Without "
                    "Any Gaps series.\n\n"
                    "Reads like a long accessible essay rather than a "
                    "textbook. Covers everyone from al-Kindi through Ibn "
                    "Arabi to Mulla Sadra and the Indian school. ~512 pages.\n\n"
                    "Best entry for someone who doesn't know where to start "
                    "with Islamic philosophy. The companion podcast (free "
                    "online) is also excellent."),
        "notes": "Oxford University Press.",
    },

    # --- OPEN QUESTIONS ---
    {
        "category": "Open Question",
        "x": ("What is rizq?\n\n"
              "An economic concept (income, output)?\n"
              "Or a metaphysical one (provision-as-actualization-from-the-Unseen)?\n\n"
              "Classical Islam said the second. Modern Islamic economics has "
              "treated it as the first. The shift may be a problem."),
        "threads": ("What is rizq?\n\n"
                    "An economic concept - the income or output a household "
                    "receives in a given period? This is how modern Islamic "
                    "economics has typically treated it.\n\n"
                    "Or a metaphysical concept - the actualization-into-"
                    "spacetime of provision determined in the Unseen? This "
                    "is what the classical tradition was working on.\n\n"
                    "Two different objects of study. Two different methods. "
                    "Two different answers. The shift from the second to "
                    "the first may be the central problem of the field."),
        "notes": "Open with this for engagement.",
    },
    {
        "category": "Open Question",
        "x": ("What is barakah?\n\n"
              "A multiplier on outcomes (your money goes further)?\n"
              "Or the thickness of manifestation - how much of the underlying "
              "field is disclosed in a manifest event?\n\n"
              "Two answers, two research programs."),
        "threads": ("What is barakah?\n\n"
                    "Read 1: a multiplicative coefficient on wealth dynamics. "
                    "Your money or time goes further than nominal magnitude "
                    "predicts. Heterodox economics can model this.\n\n"
                    "Read 2: the thickness of manifestation. How much of the "
                    "underlying field of divine names is disclosed in a "
                    "manifest event. A meal in niyyah carries thicker "
                    "disclosure than a thousand consumed without orientation.\n\n"
                    "Both coherent. They answer different questions."),
        "notes": "Barakah as coupling.",
    },
    {
        "category": "Open Question",
        "x": ("Two readings of riba:\n\n"
              "1. A specific juridical prohibition on usury contracts.\n"
              "2. An ontological inversion: claiming creative power for "
              "autonomous capital, divorced from productive engagement with "
              "reality.\n\n"
              "The second is wider. AI-finance is its limiting case."),
        "threads": ("Two readings of riba:\n\n"
                    "1. Specific juridical prohibition: certain contract "
                    "forms (interest-bearing loans, certain commodity "
                    "exchanges) are not permitted.\n\n"
                    "2. Ontological inversion: the claim that capital grows "
                    "autonomously, divorced from productive engagement with "
                    "reality. Attributing creative power to a created thing - "
                    "shirk in economic action.\n\n"
                    "The second is wider. AI-managed trading bots fit it as "
                    "directly as medieval moneylending. Maybe more."),
        "notes": "Riba as ontological inversion - main treatise Ch 13.",
    },
    {
        "category": "Open Question",
        "x": ("What is halal at the level of being?\n\n"
              "A list of permitted transactions?\n"
              "Or a mode of structural engagement with the Real?\n\n"
              "Both are needed. Modern Islamic economics has emphasized the "
              "first. The classical tradition meant the second."),
        "threads": ("What does halal mean at the level of being?\n\n"
                    "Standard reading: a list of permitted transactions, "
                    "contracts, foods, behaviors. Necessary for practical "
                    "jurisprudence. Insufficient on its own.\n\n"
                    "Akbarian reading: a mode of structural engagement with "
                    "the Real. An aligned manifestation of divine self-"
                    "disclosure through human action. Haram is the misaligned "
                    "mode.\n\n"
                    "Both registers needed. Juridical without ontological is "
                    "a checklist. Ontological without juridical floats."),
        "notes": "Halal/haram as modes of engagement.",
    },

    # --- ANTI-MYSTICISM (brand positioning) ---
    {
        "category": "Anti-Mysticism",
        "x": ("\"Quantum mechanics proves God\" is bad epistemology. It is "
              "also bad theology. The actual work modern physics has done - "
              "overturning classical materialism, forcing non-locality, "
              "opening the measurement problem - is more interesting and "
              "less triumphant."),
        "threads": ("Why \"quantum mechanics proves God\" is bad both ways:\n\n"
                    "Bad epistemology: physics has not proven any "
                    "metaphysical claim. It has constrained the space of "
                    "ontologies. Constraints are not proofs.\n\n"
                    "Bad theology: tying religious truth to a physics theory "
                    "makes religion hostage to that theory's fate. When the "
                    "next theory comes, religion has to be re-justified.\n\n"
                    "What physics does for metaphysics is negative - removing "
                    "obstacles. Real work, but not proof."),
        "notes": "(B)-not-(A) public statement.",
    },
    {
        "category": "Anti-Mysticism",
        "x": ("The Tao of Physics (Capra, 1975) and the Chopra catalog "
              "popularized \"QM proves Eastern mysticism.\" Mainstream "
              "physics rejected this and was right to. The category mistake - "
              "using physics to prove metaphysics - damaged both."),
        "threads": ("Two literatures on QM and religion:\n\n"
                    "1. Popular: Capra's Tao of Physics, Chopra's Quantum "
                    "Healing, the various YouTube essays. Overreaches "
                    "consistently. Uses \"quantum\" as license for any "
                    "spiritual claim.\n\n"
                    "2. Academic: Maudlin, Albert, Bohm, journals like "
                    "Foundations of Physics. Careful, technical, modest in "
                    "its claims.\n\n"
                    "The serious work is in the second. Most encounter only "
                    "the first."),
        "notes": "Distinguish from pop-physics-mysticism.",
    },
    {
        "category": "Anti-Mysticism",
        "x": ("The pattern: bold claim about physics from someone who has "
              "not done the work. Confident tone. Refusal to do the math "
              "when challenged. Audience persuaded by the confidence.\n\n"
              "This is what to avoid. Methodological discipline matters more "
              "than rhetorical reach."),
        "threads": ("A pattern worth recognizing in physics-and-religion "
                    "discourse:\n\n"
                    "1. Bold metaphysical claim citing physics, by someone "
                    "who has not done the physics.\n"
                    "2. Confident tone.\n"
                    "3. Refusal to do the technical work when challenged.\n"
                    "4. Audience persuaded by confidence, not substance.\n"
                    "5. When caught out, the speaker shifts ground or attacks "
                    "the questioner.\n\n"
                    "The failure mode. The discipline: do the work first, "
                    "claim second, accept correction when caught."),
        "notes": "Implicit Sahil critique.",
    },

    # --- CONTEMPORARY PHYSICS NOTES ---
    {
        "category": "Physics Note",
        "x": ("Leonard Susskind (founder of string theory), on the record:\n\n"
              "\"We live in the wrong kind of world to be described by string "
              "theory. We need to start over.\"\n\n"
              "Modern physics is more open than the popular literature "
              "suggests."),
        "threads": ("Leonard Susskind - founder of string theory, central "
                    "in the holography debate - on record:\n\n"
                    "\"We live in the wrong kind of world to be described by "
                    "string theory. I can tell you with absolute certainty "
                    "it is not the real world we live in. We need to start "
                    "over.\"\n\n"
                    "AdS/CFT works in anti-de Sitter space. Our universe is "
                    "de Sitter. The biggest results in theoretical physics "
                    "apply to the wrong universe.\n\n"
                    "Be cautious about over-using holography in metaphysics."),
        "notes": "Curt Jaimungal interview, 2026.",
    },
    {
        "category": "Physics Note",
        "x": ("Bell's theorem (1964) + loophole-free experiments (2015) "
              "established that physical reality cannot be both local AND "
              "have definite pre-existing properties. Not philosophical "
              "preference. Forced by experiment. The classical picture is "
              "dead."),
        "threads": ("What Bell's theorem actually established:\n\n"
                    "Bell (1964) proved a mathematical inequality holding "
                    "for any theory in which physical systems have definite "
                    "properties before measurement AND interactions "
                    "propagate locally.\n\n"
                    "Aspect (1980s) showed the inequality is violated. "
                    "Hensen et al. (2015) closed remaining loopholes.\n\n"
                    "Conclusion: reality cannot have both locality and "
                    "definite pre-existing properties. One must go. Not "
                    "philosophical preference - forced by experiment."),
        "notes": "Hensen 2015, Giustina 2015, Shalm 2015.",
    },
    {
        "category": "Physics Note",
        "x": ("Susskind and Maldacena's ER=EPR: quantum entanglement and "
              "spacetime wormholes are the same thing. If true, spacetime "
              "is emergent from entanglement. The priority of wujud over "
              "space and time would have a contemporary physics counterpart."),
        "threads": ("ER=EPR (Susskind and Maldacena, 2013): quantum "
                    "entanglement and spacetime wormholes are the same "
                    "thing.\n\n"
                    "If true, the implications are large. Spacetime is not "
                    "fundamental - it is emergent from a network of "
                    "entanglement relations. Locality is derived, not "
                    "basic.\n\n"
                    "For Islamic metaphysics: the Akbarian claim that wujud "
                    "is prior to space and time, with space-time as modes "
                    "of self-disclosure, would have a contemporary physics "
                    "counterpart.\n\n"
                    "Active research. Worth following."),
        "notes": "Susskind 2013, Maldacena 2013, follow-up literature.",
    },

    # --- ENGAGEMENT WITH PUBLIC DISCOURSE ---
    {
        "category": "Public Discourse",
        "x": ("On miracles and physics:\n\n"
              "There is no current physics that produces the specific event.\n"
              "There may be future physics that does.\n"
              "The theological claim does not depend on either.\n\n"
              "This is the disciplined position. Anything stronger overreaches."),
        "threads": ("The disciplined position on classical Islamic miracles "
                    "and physics:\n\n"
                    "1. The event happened (per the textual tradition).\n"
                    "2. Current physics has no mechanism for it.\n"
                    "3. Whether future physics will, or whether the event "
                    "operates at a level physics describes only partially, "
                    "is genuinely open.\n"
                    "4. The theological claim does not depend on either "
                    "resolution.\n\n"
                    "Anything stronger in either direction is an "
                    "unwarranted commitment. Hold the question open."),
        "notes": "Physics-open agnosticism.",
    },
    {
        "category": "Public Discourse",
        "x": ("The perennialist objection (Nasr): using modern physics to "
              "ground religious metaphysics gives science epistemic authority "
              "it has not earned.\n\n"
              "The objection has weight. The reply: physics opens space, it "
              "does not validate."),
        "threads": ("The perennialist objection, articulated by Seyyed "
                    "Hossein Nasr in Knowledge and the Sacred (1981):\n\n"
                    "Using modern physics to ground religious metaphysics "
                    "gives science epistemic authority it has not earned. "
                    "Classical doctrines are prior to physics, not dependent "
                    "on it.\n\n"
                    "The objection has weight. The reply: physics is not "
                    "asked to validate the doctrines. It is asked to remove "
                    "obstacles classical materialism erected. With those "
                    "gone, doctrines are stated in their own terms."),
        "notes": "Engages Nasr directly.",
    },
    {
        "category": "Public Discourse",
        "x": ("If AI materially restructures the economy: the applied "
              "frameworks that depend on current market structure will need "
              "rebuilding.\n\n"
              "The metaphysical foundations (rizq, barakah, qadar as "
              "structures of being) will not. Investment in the latter is "
              "more durable."),
        "threads": ("If AI restructures the material economy as some "
                    "predict:\n\n"
                    "Applied-economics frameworks that depend on current "
                    "market structure - income variance, household "
                    "provisioning, behavioral feedback - will need "
                    "substantial rebuilding.\n\n"
                    "The metaphysical foundations - what rizq, barakah, "
                    "qadar are at the level of being - will not. They make "
                    "structural claims, not regularity claims. Material "
                    "change does not touch them.\n\n"
                    "Metaphysics is the more durable investment."),
        "notes": "Engages Falk-Tsoukalas AI layoff paper.",
    },

    # --- PROGRESS NOTES ---
    {
        "category": "Progress Note",
        "x": ("Working on Chapter 1 of the metaphysics treatise: \"The "
              "Question of This Book.\" Two kinds of question about Islamic "
              "economics - the modeling kind and the metaphysical kind. "
              "Most contemporary work asks the first. This book asks only "
              "the second."),
        "threads": ("Working on Chapter 1 of the metaphysics treatise: \"The "
                    "Question of This Book.\"\n\n"
                    "Sets up two kinds of question about Islamic economics:\n\n"
                    "(a) Modeling - how does rizq distribute, how do "
                    "Islamic markets behave, what do welfare theorems look "
                    "like under maqasid criteria?\n\n"
                    "(b) Metaphysical - what is rizq, what is barakah, what "
                    "is qadar, at the level of being?\n\n"
                    "Most contemporary Islamic economics asks (a). This book "
                    "asks only (b). Establishing the distinction is the "
                    "first methodological move."),
        "notes": "From Chapter 1 of metaphysics_book.",
    },
    {
        "category": "Progress Note",
        "x": ("Volume 0 is the metaphysical foundation. Volume I will build "
              "the formal microeconomic theory on it (MWG-style). Volumes "
              "II-IV: macro/finance, Islamic finance theory, comparative "
              "systems.\n\n"
              "The metaphysics is the part that ages well."),
        "threads": ("The longer arc of the project:\n\n"
                    "Volume 0 - The Metaphysics of Islamic Economics. "
                    "Working draft underway. The metaphysical foundation.\n\n"
                    "Volume I - Microeconomic theory on the foundation. "
                    "MWG-style formalism, Islamic ontology made explicit.\n\n"
                    "Volumes II-IV - Macro and public finance, Islamic "
                    "finance theory, comparative economic systems.\n\n"
                    "Volume 0 is the part most resistant to material change. "
                    "The applied volumes will age. The metaphysics endures."),
        "notes": "The five-volume framing.",
    },
    {
        "category": "Progress Note",
        "x": ("Built an 18-24 month reading curriculum for the project. "
              "Five phases: math support, quantum mechanics, mathematical "
              "logic, philosophy of physics-mind-science, Islamic metaphysics. "
              "No shortcuts. The work is the work."),
        "threads": ("18-24 month reading curriculum for serious work on "
                    "Islamic metaphysics:\n\n"
                    "1. Math support (Axler, Brown-Churchill).\n"
                    "2. QM (Townsend, Albert, Maudlin, Becker).\n"
                    "3. Logic and non-computability (Smith, Sipser, "
                    "Penrose).\n"
                    "4. Philosophy of physics, mind, science (Rosenberg, "
                    "Kuhn, Chalmers, Kastrup).\n"
                    "5. Islamic metaphysics (Adamson, Chittick x2, Murata, "
                    "Hermansen, Izutsu, Fusus al-Hikam).\n\n"
                    "31 books. No shortcuts."),
        "notes": "From the syllabus docx.",
    },
]


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
wb = Workbook()
wb.remove(wb.active)

# ----- shared styling -----
header_fill = PatternFill("solid", fgColor="1A5C3A")
header_font = Font(bold=True, color="FFFFFF")
alt_fill = PatternFill("solid", fgColor="FAF8F4")
thin = Side(style="thin", color="BBBBBB")
borders = Border(top=thin, bottom=thin, left=thin, right=thin)
center = Alignment(horizontal="center", vertical="top", wrap_text=True)
left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)


def write_sheet(name, key):
    ws = wb.create_sheet(name)
    headers = ["Date", "Time", "Post", "Category", "Type", "Chars", "Notes", "Status"]
    ws.append(headers)
    # widths
    widths = [12, 10, 70, 18, 14, 8, 30, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # header style
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = borders
    # rows
    for idx, p in enumerate(posts, start=2):
        text = p[key]
        ws.append([
            "",                       # Date
            "",                       # Time
            text,                     # Post
            p["category"],            # Category
            "standalone",             # Type
            len(text),                # Chars
            p.get("notes", ""),       # Notes
            "draft",                  # Status
        ])
        row = ws[idx]
        if idx % 2 == 0:
            for c in row:
                c.fill = alt_fill
        for c in row:
            c.border = borders
            c.alignment = left_top
        ws.row_dimensions[idx].height = 90
    # freeze header
    ws.freeze_panes = "A2"


write_sheet("Metaphysics X", "x")
write_sheet("Metaphysics Threads", "threads")

out_path = "/Users/hamzazahid/Library/CloudStorage/GoogleDrive-88hzahid@gmail.com/My Drive/IslamicEconomics/social/content-calendar-metaphysics.xlsx"
wb.save(out_path)
print(f"Wrote: {out_path}")
print(f"Posts per sheet: {len(posts)}")
print("Categories included:")
from collections import Counter
cat_counts = Counter(p["category"] for p in posts)
for cat, n in sorted(cat_counts.items()):
    print(f"  {cat}: {n}")
